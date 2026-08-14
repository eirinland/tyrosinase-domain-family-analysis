"""
Network graph of active-site vectors, v2.
Nodes = unique vectors (sized by member count, min threshold applied).
Edges = Hamming distance <= threshold (~ at Gly46 treated as wildcard).
Colored by characterized activity + marker-defined pools in lighter shades.
"""

import csv
from pathlib import Path
from collections import Counter

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.colors import to_rgb
import networkx as nx
import openpyxl

BASE = Path(__file__).parent.parent
VECTORS_CSV = BASE / 'position_vectors.csv'
CHAR_XLSX = Path('/cluster/work/projects/nn1003k/eirin/bioinf/characterized_PPOs.xlsx')

MIN_SIZE = 3
HAMMING_THRESHOLD = 1

ACTIVITY_COLOR = {
    'TYR':        '#4C8CC2',
    'CaOx':       '#87B5D1',
    'AUS':        '#9ECCC7',
    'oAPO':       '#A6D1A6',
    'oMP':        '#6BAB76',
    'DHICA ox':   '#F0CFB0',
    'DCT':        '#D9A68A',
    'hemocyanin': '#B59EC7',
}

POOL_MARKERS = {
    'oAPO':       ('Gly46', 'N'),
    'oMP':        ('Arg209', 'Y'),
    'DCT/DHICA':  ('His230', 'L'),
    'hemocyanin': ('Gly46', 'E'),
}

POOL_COLOR = {
    'oAPO':       '#D3E8D3',
    'oMP':        '#B5D5BA',
    'DCT/DHICA':  '#F5E4D5',
    'hemocyanin': '#DACFE5',
}

COL = {
    'acc': 0, 'gly46': 6, 'gly46_ss': 8, 'phe65': 14, 'trp68': 18,
    'glu195': 26, 'asn205': 34, 'arg209': 42, 'val218': 46, 'ala221': 50,
    'phe227': 54, 'his230': 58, 'thioether': 66, 'vector': 68,
}

POS_COLS = {
    'Gly46': 6, 'Phe65': 14, 'Trp68': 18, 'Glu195': 26, 'Asn205': 34,
    'Arg209': 42, 'Val218': 46, 'Ala221': 50, 'Phe227': 54, 'His230': 58,
}


def hamming(v1, v2):
    p1 = v1.split('-')
    p2 = v2.split('-')
    d = 0
    for a, b in zip(p1, p2):
        if a == '~' or b == '~':
            continue
        if a != b:
            d += 1
    return d


def load_characterized():
    wb = openpyxl.load_workbook(CHAR_XLSX, read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    acc_to_act = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc = str(row[0]).strip()
        act = str(row[1]).strip()
        acc_to_act[acc] = act
    wb.close()
    return acc_to_act


def main():
    char_map = load_characterized()

    vec_count = Counter()
    vec_accs = {}
    vec_activities = {}
    vec_pools = {}
    acc_to_vec = {}

    with open(VECTORS_CSV) as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            acc = row[COL['acc']]
            vec = row[COL['vector']].replace('*', '')
            vec_count[vec] += 1

            if vec not in vec_accs:
                vec_accs[vec] = []
            vec_accs[vec].append(acc)
            acc_to_vec[acc] = vec

            if acc in char_map:
                if vec not in vec_activities:
                    vec_activities[vec] = set()
                vec_activities[vec].add(char_map[acc])

            gly46 = row[COL['gly46']].replace('*', '')
            arg209 = row[COL['arg209']].replace('*', '')
            his230 = row[COL['his230']].replace('*', '')
            if vec not in vec_pools:
                vec_pools[vec] = set()
            if gly46 == 'N':
                vec_pools[vec].add('oAPO')
            if arg209 == 'Y':
                vec_pools[vec].add('oMP')
            if his230 == 'L':
                vec_pools[vec].add('DCT/DHICA')
            if gly46 == 'E':
                vec_pools[vec].add('hemocyanin')

    keep = {v: c for v, c in vec_count.items() if c >= MIN_SIZE}
    vectors = sorted(keep.keys())
    print(f"Vectors with >= {MIN_SIZE} members: {len(vectors)}")
    print(f"Characterized vectors in graph: "
          f"{sum(1 for v in vectors if v in vec_activities)}")

    G = nx.Graph()
    for v in vectors:
        G.add_node(v, count=keep[v])

    n = len(vectors)
    edges_by_hd = {d: [] for d in range(1, HAMMING_THRESHOLD + 1)}
    for i in range(n):
        for j in range(i + 1, n):
            d = hamming(vectors[i], vectors[j])
            if 1 <= d <= HAMMING_THRESHOLD:
                G.add_edge(vectors[i], vectors[j], hamming=d,
                           weight=max(HAMMING_THRESHOLD + 1 - d, 0.5))
                edges_by_hd[d].append((vectors[i], vectors[j]))

    total_edges = sum(len(e) for e in edges_by_hd.values())
    print(f"Edges (Hamming <= {HAMMING_THRESHOLD}): {total_edges}")
    for d in sorted(edges_by_hd):
        print(f"  HD={d}: {len(edges_by_hd[d])}")

    components = list(nx.connected_components(G))
    print(f"Connected components: {len(components)}")
    print(f"Largest component: {max(len(c) for c in components)} nodes")

    pos = nx.spring_layout(G, k=2.0, iterations=200, seed=42, weight='weight')

    sizes = np.array([keep[v] for v in vectors])
    log_sizes = np.log10(np.clip(sizes, 1, None))
    rng = max(log_sizes.max() - log_sizes.min(), 1)
    pt_sizes = 15 + (log_sizes - log_sizes.min()) / rng * 300

    # Assign colors: characterized > pool > gray
    colors = []
    for v in vectors:
        if v in vec_activities:
            acts = sorted(vec_activities[v])
            colors.append(ACTIVITY_COLOR.get(acts[0], '#555555'))
        elif v in vec_pools and vec_pools[v]:
            pools = sorted(vec_pools[v])
            colors.append(POOL_COLOR.get(pools[0], '#dddddd'))
        else:
            colors.append('#dddddd')

    fig, ax = plt.subplots(figsize=(14, 12))

    edge_styles = [
        (1, '#aaaaaa', 0.5, 0.5, 'solid'),
    ]
    for d, color, width, alpha, style in edge_styles:
        if edges_by_hd.get(d):
            nx.draw_networkx_edges(G, pos, edgelist=edges_by_hd[d], ax=ax,
                                   edge_color=color, width=width, alpha=alpha,
                                   style=style)

    # Draw uncharacterized, non-pool nodes
    plain_idx = [i for i, v in enumerate(vectors)
                 if v not in vec_activities and not (v in vec_pools and vec_pools[v])]
    if plain_idx:
        xy = np.array([pos[vectors[i]] for i in plain_idx])
        ax.scatter(xy[:, 0], xy[:, 1],
                   s=[pt_sizes[i] for i in plain_idx],
                   c=[colors[i] for i in plain_idx],
                   edgecolors='#aaaaaa', linewidths=0.2, alpha=0.5, zorder=2)

    # Draw pool nodes (not characterized)
    pool_idx = [i for i, v in enumerate(vectors)
                if v not in vec_activities and v in vec_pools and vec_pools[v]]
    if pool_idx:
        xy = np.array([pos[vectors[i]] for i in pool_idx])
        ax.scatter(xy[:, 0], xy[:, 1],
                   s=[pt_sizes[i] for i in pool_idx],
                   c=[colors[i] for i in pool_idx],
                   edgecolors='#888888', linewidths=0.4, alpha=0.7, zorder=3)

    # Draw characterized nodes on top
    char_idx = [i for i, v in enumerate(vectors) if v in vec_activities]
    if char_idx:
        xy = np.array([pos[vectors[i]] for i in char_idx])
        ax.scatter(xy[:, 0], xy[:, 1],
                   s=[pt_sizes[i] * 2.5 for i in char_idx],
                   c=[colors[i] for i in char_idx],
                   edgecolors='black', linewidths=1.5, alpha=1.0, zorder=4,
                   marker='D')

    # Legend
    legend_elements = []
    for act in ['TYR', 'CaOx', 'AUS', 'oAPO', 'oMP', 'DHICA ox', 'DCT', 'hemocyanin']:
        if any(act in vec_activities.get(v, set()) for v in vectors):
            legend_elements.append(
                Line2D([0], [0], marker='D', color='w',
                       markerfacecolor=ACTIVITY_COLOR[act],
                       markeredgecolor='black', markersize=9, label=f'{act} (characterized)'))

    legend_elements.append(Line2D([0], [0], marker='s', color='w', markersize=0, label=''))
    for pool_name in ['oAPO', 'oMP', 'DCT/DHICA', 'hemocyanin']:
        legend_elements.append(
            Line2D([0], [0], marker='o', color='w',
                   markerfacecolor=POOL_COLOR[pool_name],
                   markeredgecolor='#888888', markersize=8,
                   label=f'{pool_name} pool'))

    legend_elements.append(
        Line2D([0], [0], marker='o', color='w',
               markerfacecolor='#dddddd', markeredgecolor='#aaaaaa',
               markersize=7, label='Other'))

    legend_elements.append(Line2D([0], [0], marker='s', color='w', markersize=0, label=''))
    for s_example in [5, 50, 500]:
        if s_example <= sizes.max():
            ls = np.log10(s_example)
            ps = 15 + (ls - log_sizes.min()) / rng * 300
            legend_elements.append(
                Line2D([0], [0], marker='o', color='w', markerfacecolor='#aaaaaa',
                       markersize=np.sqrt(ps) * 0.55, label=f'n = {s_example}'))

    ax.legend(handles=legend_elements, loc='upper left', fontsize=8.5,
              framealpha=0.95, title='Activity / Pool / Size',
              title_fontsize=9, borderpad=1)

    ax.set_xticks([])
    ax.set_yticks([])
    ax.axis('off')

    out = Path(__file__).parent / 'network_graph_v2'
    fig.savefig(f'{out}.pdf', dpi=300, bbox_inches='tight')
    fig.savefig(f'{out}.png', dpi=200, bbox_inches='tight')
    plt.close()
    print(f"Saved: {out}.pdf / .png")


if __name__ == '__main__':
    main()
