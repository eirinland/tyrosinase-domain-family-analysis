"""
Activity controller analysis for the PPO manuscript.
Reads characterized_PPOs.xlsx, position_vectors.csv, all_vectors.csv,
visualisation/taxonomy_lookup.csv, and the FoldSeek cluster file.
Prints all results to stdout.
"""

import csv
import sys
from pathlib import Path
from collections import Counter, defaultdict
from itertools import combinations

import openpyxl

HERE = Path(__file__).parent
XLSX = HERE / '../data/characterized_PPOs.xlsx'
PV_CSV = HERE / 'position_vectors.csv'
AV_CSV = HERE / 'all_vectors.csv'
TAX_CSV = HERE / 'visualisation/taxonomy_lookup.csv'
CLUSTER_TSV = HERE / '../5_structural_clustering/results/cluster_cluster.tsv'

POSITIONS = ['Gly46', 'Phe65', 'Trp68', 'Glu195', 'Asn205',
             'Arg209', 'Val218', 'Ala221', 'Phe227', 'His230', 'thioether']
CANONICAL = {'Gly46': 'G', 'Phe65': 'F', 'Trp68': 'W', 'Glu195': 'E',
              'Asn205': 'N', 'Arg209': 'R', 'Val218': 'V', 'Ala221': 'A',
              'Phe227': 'F', 'His230': 'H', 'thioether': 'C'}


def norm_res(r):
    """Strip asterisk uncertainty marker and normalise loop chars."""
    if r is None:
        return '?'
    r = str(r).strip().rstrip('*')
    return '@' if r == '~' else r


def norm_vec(v):
    return v.replace('~', '@') if v else v


def vec_parts(v):
    return [norm_res(p) for p in v.split('-')]


def hamming(a, b):
    return sum(1 for x, y in zip(a, b) if norm_res(x) != norm_res(y))


def sep_print(title=''):
    print()
    print('=' * 70)
    if title:
        print(title)
        print('=' * 70)


# ── Load data ──────────────────────────────────────────────────────────────

def load_xlsx():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        row = {h: (str(v).strip() if v is not None else '') for h, v in zip(headers, r)}
        rows.append(row)
    return rows


def load_position_vectors():
    rows = []
    with open(PV_CSV) as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows


def load_all_vectors():
    rows = []
    with open(AV_CSV) as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows


def load_taxonomy():
    tax = {}
    with open(TAX_CSV) as f:
        for row in csv.DictReader(f):
            tax[row['accession']] = row
    return tax


def load_clusters():
    """Returns {member: rep} and {rep: [members]}."""
    member2rep = {}
    rep2members = defaultdict(list)
    with open(CLUSTER_TSV) as f:
        for line in f:
            parts = line.strip().split('\t')
            if len(parts) < 2:
                continue
            rep, member = parts[0], parts[1]
            member2rep[member] = rep
            rep2members[rep].append(member)
    return member2rep, rep2members


def pv_res(row, pos):
    """Get residue from position_vectors row, applying loop rule."""
    if pos == 'thioether':
        v = row.get('thioether', '-')
        return 'C' if v in ('C', 'C*') else '-'
    res = norm_res(row.get(f'{pos}_res', row.get(pos, '?')))
    if pos == 'Gly46' and row.get('Gly46_ss', '') == 'c':
        return '@'
    return res


def xlsx_res(row, pos):
    """Get residue from characterized_PPOs xlsx row."""
    val = norm_res(row.get(pos, '?'))
    if val == '':
        return '?'
    return val


def build_sig(row_getter, row, positions):
    return tuple(row_getter(row, p) for p in positions)


def build_vec_from_pv(row):
    """Build full 11-position vector tuple from position_vectors row."""
    return tuple(pv_res(row, p) for p in POSITIONS)


def build_vec_from_xlsx(row):
    return tuple(xlsx_res(row, p) for p in POSITIONS)


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    print("Loading data...", file=sys.stderr)
    char_rows = load_xlsx()
    pv_rows = load_position_vectors()
    av_rows = load_all_vectors()
    tax = load_taxonomy()
    member2rep, rep2members = load_clusters()
    print(f"  Characterized: {len(char_rows)}", file=sys.stderr)
    print(f"  All structures: {len(pv_rows)}", file=sys.stderr)

    # Index characterized by accession
    char_by_acc = {r['Accession']: r for r in char_rows}

    # Build activity sets
    tyr_rows = [r for r in char_rows if r.get('Activity') == 'TYR']
    caox_rows = [r for r in char_rows if r.get('Activity') == 'CaOx']

    # Build pv lookup by accession
    pv_by_acc = {r['accession']: r for r in pv_rows}

    # ── ANALYSIS 1 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 1: Single-position TYR vs CaOx separation')
    print(f'\nTYR n={len(tyr_rows)}, CaOx n={len(caox_rows)}')

    invariant = []
    variable = []
    for pos in POSITIONS:
        tyr_vals = Counter(xlsx_res(r, pos) for r in tyr_rows)
        caox_vals = Counter(xlsx_res(r, pos) for r in caox_rows)
        tyr_uniq = set(tyr_vals) - {'?'}
        caox_uniq = set(caox_vals) - {'?'}
        if len(tyr_uniq) == 1 and len(caox_uniq) == 1:
            invariant.append((pos, list(tyr_uniq)[0], list(caox_uniq)[0]))
        else:
            variable.append((pos, tyr_vals, caox_vals))

    print('\nInvariant positions (same residue in all TYR and all CaOx):')
    for pos, tval, cval in invariant:
        same = 'SAME' if tval == cval else 'DIFFERENT'
        print(f'  {pos}: TYR={tval}, CaOx={cval} [{same}]')

    print('\nVariable positions:')
    for pos, tyr_vals, caox_vals in variable:
        tyr_str = ', '.join(f'{r}:{n}' for r, n in sorted(tyr_vals.items(), key=lambda x: -x[1]) if r != '?')
        caox_str = ', '.join(f'{r}:{n}' for r, n in sorted(caox_vals.items(), key=lambda x: -x[1]) if r != '?')
        print(f'  {pos}:')
        print(f'    TYR:  {tyr_str}')
        print(f'    CaOx: {caox_str}')

    # Gatekeeper positions: Trp68 and Val218
    print('\nGatekeeper analysis (Trp68 and Val218):')
    for pos, caox_typical in [('Trp68', 'F'), ('Val218', 'A')]:
        tyr_with_caox = [r for r in tyr_rows if xlsx_res(r, pos) == caox_typical]
        print(f'\n  TYR entries with {pos}={caox_typical} (CaOx-typical): n={len(tyr_with_caox)}')
        for r in tyr_with_caox:
            vec = norm_vec(r.get('Vector', ''))
            print(f'    {r["Accession"]}  {r.get("Activity","?")}  vector={vec}')

    # ── ANALYSIS 2 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 2: Combinatorial TYR vs CaOx separation')
    combined = tyr_rows + caox_rows
    n_total = len(combined)

    def separation_accuracy(pos_list):
        """Fraction of entries correctly separated (unique activity per sub-sig)."""
        sig_acts = defaultdict(set)
        for r in combined:
            sig = tuple(xlsx_res(r, p) for p in pos_list)
            sig_acts[sig].add(r.get('Activity'))
        correct = sum(1 for r in combined
                      if len(sig_acts[tuple(xlsx_res(r, p) for p in pos_list)]) == 1)
        return correct / n_total

    print(f'\nAll pairs with >= 80% separation accuracy ({n_total} entries):')
    good_pairs = []
    for p1, p2 in combinations(POSITIONS, 2):
        acc = separation_accuracy([p1, p2])
        if acc >= 0.80:
            good_pairs.append((acc, p1, p2))
            print(f'  {p1} + {p2}: {acc:.1%}')

    print(f'\nTriples achieving 100% separation:')
    perfect_triples = []
    for p1, p2, p3 in combinations(POSITIONS, 3):
        acc = separation_accuracy([p1, p2, p3])
        if acc == 1.0:
            perfect_triples.append((p1, p2, p3))
            print(f'  {p1} + {p2} + {p3}')

    print(f'\nTotal perfect triples: {len(perfect_triples)}')
    # Check which contain Asn205+Arg209
    n205_209 = sum(1 for t in perfect_triples if 'Asn205' in t and 'Arg209' in t)
    print(f'Perfect triples containing Asn205+Arg209: {n205_209}')

    # ── ANALYSIS 3 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 3: Multi-activity Asn205+Arg209 signature analysis')
    activities = sorted(set(r.get('Activity', '') for r in char_rows if r.get('Activity')))
    print(f'\nActivities: {activities}')

    sig_to_acts = defaultdict(set)
    act_sigs = defaultdict(Counter)
    for r in char_rows:
        act = r.get('Activity', '')
        sig = (xlsx_res(r, 'Asn205'), xlsx_res(r, 'Arg209'))
        sig_to_acts[sig].add(act)
        act_sigs[act][sig] += 1

    print('\nAsn205+Arg209 signatures per activity:')
    for act in activities:
        sigs = act_sigs[act]
        print(f'  {act} (n={sum(sigs.values())}):')
        for sig, cnt in sorted(sigs.items(), key=lambda x: -x[1]):
            shared = len(sig_to_acts[sig]) > 1
            flag = ' [shared]' if shared else ''
            print(f'    {sig[0]}/{sig[1]}: {cnt}{flag}')

    shared_sigs = {s for s, acts in sig_to_acts.items() if len(acts) > 1}
    unique_sigs = {s for s, acts in sig_to_acts.items() if len(acts) == 1}
    print(f'\nShared signatures (occur in >1 activity): {len(shared_sigs)}')
    for s in sorted(shared_sigs):
        acts = sorted(sig_to_acts[s])
        print(f'  {s[0]}/{s[1]}: {", ".join(acts)}')
    print(f'Unique signatures (one activity only): {len(unique_sigs)}')

    # ── ANALYSIS 4 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 4: TYR entries with Phe at Trp68')
    tyr_phe68 = [r for r in tyr_rows if xlsx_res(r, 'Trp68') == 'F']
    print(f'\nTYR entries with Trp68=F (gatekeeper substitution): n={len(tyr_phe68)}')
    for r in tyr_phe68:
        vec = norm_vec(r.get('Vector', ''))
        print(f'  {r["Accession"]}  vector={vec}')
    print('Note: all are plant PPOs')

    # ── ANALYSIS 5 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 5: Phe65 and Phe227 conservation analysis')
    total_pv = len(pv_rows)
    for pos in ['Phe65', 'Phe227']:
        canon = CANONICAL[pos]
        vals = Counter(pv_res(r, pos) for r in pv_rows)
        n_canon = vals.get(canon, 0)
        print(f'\n{pos} (canonical={canon}):')
        print(f'  Conserved: {n_canon}/{total_pv} ({n_canon/total_pv:.1%})')
        subs = {aa: n for aa, n in vals.items() if aa not in (canon, '?')}
        if subs:
            print(f'  Non-canonical substitutions:')
            for aa, n in sorted(subs.items(), key=lambda x: -x[1]):
                # Taxonomic enrichment
                tax_counter = Counter()
                for r in pv_rows:
                    if pv_res(r, pos) == aa:
                        t = tax.get(r['accession'], {})
                        phylum = t.get('phylum', '?')
                        tax_counter[phylum] += 1
                top_tax = tax_counter.most_common(3)
                tax_str = ', '.join(f'{p}:{c}' for p, c in top_tax)
                print(f'    {aa}: {n} ({n/total_pv:.2%}) — {tax_str}')
        else:
            print('  No non-canonical substitutions.')

    # ── ANALYSIS 6 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 6: Chaetothyriales enrichment')
    chaeto_accs = set()
    for acc, t in tax.items():
        if 'Chaetothyriales' in t.get('phylum', '') or 'Chaetothyriales' in t.get('genus', ''):
            chaeto_accs.add(acc)
    # Also check if any contain Chaetothyriales in taxon columns
    for acc, t in tax.items():
        for col in ('phylum', 'genus'):
            if 'Chaetothyriales' in str(t.get(col, '')):
                chaeto_accs.add(acc)

    pv_accs = set(r['accession'] for r in pv_rows)
    chaeto_in_ds = chaeto_accs & pv_accs
    print(f'\nChaetothyriales structures in dataset: {len(chaeto_in_ds)}/{total_pv}')
    bg_rate = len(chaeto_in_ds) / total_pv if total_pv > 0 else 0
    print(f'Background rate: {bg_rate:.3%}')

    print('\nEnrichment in non-canonical substitutions:')
    for pos in POSITIONS:
        if pos == 'thioether':
            continue
        canon = CANONICAL.get(pos, '?')
        non_canon_accs = [r['accession'] for r in pv_rows if pv_res(r, pos) not in (canon, '?', '@')]
        if not non_canon_accs:
            continue
        n_non = len(non_canon_accs)
        n_chaeto = sum(1 for a in non_canon_accs if a in chaeto_in_ds)
        rate = n_chaeto / n_non if n_non > 0 else 0
        fold = rate / bg_rate if bg_rate > 0 else 0
        if fold > 2 or n_chaeto > 5:
            print(f'  {pos}: non-canonical n={n_non}, Chaetothyriales={n_chaeto} ({rate:.1%}), '
                  f'enrichment={fold:.1f}x')

    # ── ANALYSIS 7 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 7: Oomycota cluster 5 and Gly46')
    # Identify clusters by size and kingdom composition
    # Enumerate clusters by descending member count to find cluster 5
    cluster_sizes = Counter(member2rep.get(r['accession'], r['accession']) for r in pv_rows)
    # Reps sorted by size
    top_reps = [rep for rep, _ in cluster_sizes.most_common()]

    if len(top_reps) >= 5:
        cluster5_rep = top_reps[4]
        cluster5_members = set(rep2members.get(cluster5_rep, [cluster5_rep]))
        # Filter to those in pv_rows
        cluster5_pv = [r for r in pv_rows if r['accession'] in cluster5_members]

        # Kingdom breakdown of cluster 5
        k_cnt = Counter()
        for r in cluster5_pv:
            t = tax.get(r['accession'], {})
            k_cnt[t.get('kingdom', '?')] += 1
        # Phylum breakdown
        ph_cnt = Counter()
        for r in cluster5_pv:
            t = tax.get(r['accession'], {})
            ph_cnt[t.get('phylum', '?')] += 1

        print(f'\nCluster 5 (rep={cluster5_rep}): {len(cluster5_pv)} structures in pv dataset')
        print(f'Kingdom: {dict(k_cnt.most_common())}')
        print(f'Top phyla: {ph_cnt.most_common(5)}')

        print('\nGly46 distribution in cluster 5:')
        gly46_cnt = Counter(pv_res(r, 'Gly46') for r in cluster5_pv)
        for res, n in sorted(gly46_cnt.items(), key=lambda x: -x[1]):
            # SS breakdown for this residue
            ss_cnt = Counter()
            for r in cluster5_pv:
                if pv_res(r, 'Gly46') == res:
                    ss_cnt[r.get('Gly46_ss', '?')] += 1
            print(f'  {res}: {n} — SS: {dict(ss_cnt)}')
    else:
        print(f'\nFewer than 5 clusters found ({len(top_reps)} total). Showing all.')
        for i, rep in enumerate(top_reps[:10]):
            members = rep2members.get(rep, [rep])
            pv_members = [r for r in pv_rows if r['accession'] in set(members)]
            k_cnt = Counter(tax.get(r['accession'], {}).get('kingdom', '?') for r in pv_members)
            print(f'  Cluster {i+1}: {len(pv_members)} pv structures, kingdoms={dict(k_cnt.most_common(3))}')

    # ── ANALYSIS 8 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 8: DHICA ox, DCT, vertebrate TYR in structural network')
    special_acts = ['DHICA ox', 'DCT', 'TYR']
    for act in special_acts:
        act_char = [r for r in char_rows if r.get('Activity') == act]
        if not act_char:
            continue
        vecs = Counter(norm_vec(r.get('Vector', '')) for r in act_char if r.get('Vector'))
        print(f'\n{act} (n={len(act_char)}) vectors:')
        for v, cnt in vecs.most_common():
            # Count in full pv dataset
            pv_cnt = sum(1 for r in pv_rows if norm_vec(r.get('vector', '')) == v)
            # FoldSeek cluster
            exemplars = [r['accession'] for r in pv_rows if norm_vec(r.get('vector', '')) == v]
            if exemplars:
                rep = member2rep.get(exemplars[0], exemplars[0])
                cluster_rank = (top_reps.index(rep) + 1) if rep in top_reps else '?'
            else:
                cluster_rank = '?'
            print(f'  vector={v} | characterized={cnt} | pv_total={pv_cnt} | cluster_rank={cluster_rank}')

    # ── ANALYSIS 9 ─────────────────────────────────────────────────────────
    sep_print('ANALYSIS 9: Full dataset partitioning by Asn205+Arg209')

    # Build characterized Asn205+Arg209 -> set of activities
    char_sig_acts = defaultdict(set)
    for r in char_rows:
        sig = (xlsx_res(r, 'Asn205'), xlsx_res(r, 'Arg209'))
        char_sig_acts[sig].add(r.get('Activity', '?'))

    novel_list = []
    unambig = defaultdict(list)
    ambig_list = []

    for r in pv_rows:
        sig = (pv_res(r, 'Asn205'), pv_res(r, 'Arg209'))
        matched_acts = char_sig_acts.get(sig, set())
        if not matched_acts:
            novel_list.append(r)
        elif len(matched_acts) == 1:
            act = list(matched_acts)[0]
            unambig[act].append(r)
        else:
            ambig_list.append(r)

    print(f'\nPartitioning by Asn205+Arg209 signatures (n={total_pv}):')
    n_novel = len(novel_list)
    n_ambig = len(ambig_list)
    n_unambig = sum(len(v) for v in unambig.values())
    print(f'  Novel (no match): {n_novel} ({n_novel/total_pv:.1%})')
    print(f'  Unambiguous:      {n_unambig} ({n_unambig/total_pv:.1%})')
    print(f'  Ambiguous:        {n_ambig} ({n_ambig/total_pv:.1%})')

    print('\nUnambiguous breakdown by activity:')
    for act in sorted(unambig):
        rs = unambig[act]
        k_cnt = Counter(tax.get(r['accession'], {}).get('kingdom', '?') for r in rs)
        print(f'  {act}: {len(rs)} — {dict(k_cnt.most_common(3))}')

    print('\nNovel: kingdom breakdown:')
    k_cnt = Counter(tax.get(r['accession'], {}).get('kingdom', '?') for r in novel_list)
    for k, n in k_cnt.most_common():
        print(f'  {k}: {n}')

    print('\nTop 5 novel Asn205+Arg209 signatures:')
    novel_sig_cnt = Counter()
    for r in novel_list:
        sig = (pv_res(r, 'Asn205'), pv_res(r, 'Arg209'))
        novel_sig_cnt[sig] += 1
    for sig, n in novel_sig_cnt.most_common(5):
        print(f'  {sig[0]}/{sig[1]}: {n}')

    # ── ANALYSIS 10 ────────────────────────────────────────────────────────
    sep_print('ANALYSIS 10: Full vector separation (85 characterized entries)')
    char_vecs = []
    for r in char_rows:
        v = norm_vec(r.get('Vector', ''))
        act = r.get('Activity', '?')
        char_vecs.append((v, act))

    vec_acts = defaultdict(set)
    for v, act in char_vecs:
        vec_acts[v].add(act)

    unique_vecs = len(vec_acts)
    shared_vecs = sum(1 for acts in vec_acts.values() if len(acts) > 1)
    unique_to_one = sum(1 for acts in vec_acts.values() if len(acts) == 1)

    print(f'\nUnique vectors in 85 characterized entries: {unique_vecs}')
    print(f'  Unique to one activity: {unique_to_one}')
    print(f'  Shared between activities: {shared_vecs}')

    # Within-activity convergence
    print('\nWithin-activity convergence (entries sharing same vector):')
    act_vec_cnt = defaultdict(Counter)
    for v, act in char_vecs:
        act_vec_cnt[act][v] += 1
    for act in sorted(act_vec_cnt):
        cnt = act_vec_cnt[act]
        total_act = sum(cnt.values())
        shared_in_act = sum(n for n in cnt.values() if n > 1)
        print(f'  {act}: {total_act} entries, {len(cnt)} unique vectors, '
              f'{shared_in_act} entries share a vector with another entry')

    # Cross-activity Hamming distance 1 pairs
    print('\nCross-activity pairs at Hamming distance 1:')
    all_vecs_list = [(v, sorted(acts)[0]) for v, acts in vec_acts.items() if len(acts) == 1]
    hd1_pairs = []
    for i in range(len(all_vecs_list)):
        for j in range(i + 1, len(all_vecs_list)):
            vi, ai = all_vecs_list[i]
            vj, aj = all_vecs_list[j]
            if ai != aj:
                pi, pj = vec_parts(vi), vec_parts(vj)
                if len(pi) == len(pj) and hamming(pi, pj) == 1:
                    hd1_pairs.append((vi, ai, vj, aj))
    for vi, ai, vj, aj in hd1_pairs[:20]:
        print(f'  {ai}: {vi}  ↔  {aj}: {vj}')
    if len(hd1_pairs) > 20:
        print(f'  ... ({len(hd1_pairs)} total)')

    # ── ANALYSIS 11 ────────────────────────────────────────────────────────
    sep_print('ANALYSIS 11: Position subset analysis (2047 subsets)')

    def subset_score(pos_list):
        sig_acts = defaultdict(set)
        for r in char_rows:
            sig = tuple(xlsx_res(r, p) for p in pos_list)
            sig_acts[sig].add(r.get('Activity', '?'))
        n_uniq = sum(1 for r in char_rows
                     if len(sig_acts[tuple(xlsx_res(r, p) for p in pos_list)]) == 1)
        return n_uniq / len(char_rows)

    best_by_size = {}
    for size in range(1, 12):
        best_score, best_subset = 0.0, None
        for subset in combinations(POSITIONS, size):
            sc = subset_score(list(subset))
            if sc > best_score:
                best_score = sc
                best_subset = subset
        best_by_size[size] = (best_score, best_subset)

    print(f'\nBest subset at each size (score = fraction of 85 entries with unique-activity sub-sig):')
    for size in range(1, 12):
        sc, sub = best_by_size[size]
        print(f'  Size {size:2d}: {sc:.3f}  {" + ".join(sub)}')

    print('\nTop pairs:')
    pair_scores = []
    for subset in combinations(POSITIONS, 2):
        pair_scores.append((subset_score(list(subset)), subset))
    for sc, sub in sorted(pair_scores, reverse=True)[:5]:
        print(f'  {sc:.3f}  {" + ".join(sub)}')

    print('\nTop triples:')
    triple_scores = []
    for subset in combinations(POSITIONS, 3):
        triple_scores.append((subset_score(list(subset)), subset))
    for sc, sub in sorted(triple_scores, reverse=True)[:5]:
        print(f'  {sc:.3f}  {" + ".join(sub)}')

    # ── ANALYSIS 12 ────────────────────────────────────────────────────────
    sep_print('ANALYSIS 12: Activity-specific markers and candidate pools')

    markers = [
        ('oMP marker', 'Arg209', 'Y'),
        ('oAPO marker', 'Gly46', 'N'),
        ('DCT/DHICA marker', 'His230', 'L'),
        ('hemocyanin marker', 'Gly46', 'E'),
        ('biosynthetic/oMP marker', 'Glu195', 'not_E'),
    ]

    for label, pos, target in markers:
        if target == 'not_E':
            pool = [r for r in pv_rows if pv_res(r, pos) not in ('E', '?')]
        else:
            pool = [r for r in pv_rows if pv_res(r, pos) == target]
        k_cnt = Counter(tax.get(r['accession'], {}).get('kingdom', '?') for r in pool)
        # Characterized entries matching this marker
        if target == 'not_E':
            char_match = [r for r in char_rows if xlsx_res(r, pos) not in ('E', '?')]
        else:
            char_match = [r for r in char_rows if xlsx_res(r, pos) == target]
        char_acts = Counter(r.get('Activity', '?') for r in char_match)
        print(f'\n{label} ({pos}={target}):')
        print(f'  Full dataset: {len(pool)} structures')
        print(f'  Kingdom breakdown: {dict(k_cnt.most_common(4))}')
        print(f'  Characterized entries: {len(char_match)} — {dict(char_acts)}')

    # ── ANALYSIS 13 ────────────────────────────────────────────────────────
    sep_print('ANALYSIS 13: Gly46 loop encoding limitations')

    n_loop = sum(1 for r in pv_rows if r.get('Gly46_ss', '') == 'c')
    n_resolved = total_pv - n_loop
    print(f'\nGly46 encoding:')
    print(f'  Loop (~/@): {n_loop} ({n_loop/total_pv:.1%})')
    print(f'  Resolved residue: {n_resolved} ({n_resolved/total_pv:.1%})')

    # Partitioning with 2, 3, 4, full positions
    for pos_subset, label in [
        (['Asn205', 'Arg209'], '2-pos (Asn205+Arg209)'),
        (['Gly46', 'Asn205', 'Arg209'], '3-pos (+Gly46)'),
        (['Gly46', 'Trp68', 'Asn205', 'Arg209'], '4-pos (+Trp68)'),
        (POSITIONS, 'Full 11-pos'),
    ]:
        char_sig_acts2 = defaultdict(set)
        for r in char_rows:
            char_sig_acts2[tuple(xlsx_res(r, p) for p in pos_subset)].add(r.get('Activity', '?'))

        n_novel2 = sum(1 for r in pv_rows
                       if not char_sig_acts2.get(tuple(pv_res(r, p) for p in pos_subset)))
        n_ambig2 = sum(1 for r in pv_rows
                       if len(char_sig_acts2.get(tuple(pv_res(r, p) for p in pos_subset), set())) > 1)
        n_unambig2 = total_pv - n_novel2 - n_ambig2
        print(f'\n  {label}: novel={n_novel2} ({n_novel2/total_pv:.1%}), '
              f'unambig={n_unambig2} ({n_unambig2/total_pv:.1%}), '
              f'ambig={n_ambig2} ({n_ambig2/total_pv:.1%})')

    # ── ANALYSIS 14 ────────────────────────────────────────────────────────
    sep_print('ANALYSIS 14: Hamming distance within candidate pools')

    marker_pools = [
        ('oMP', 'Arg209', 'Y', 'oMP'),
        ('oAPO', 'Gly46', 'N', 'oAPO'),
        ('DCT/DHICA', 'His230', 'L', ['DCT', 'DHICA ox']),
        ('hemocyanin', 'Gly46', 'E', 'hemocyanin'),
    ]

    # Build tuples for characterized by activity
    char_vecs_by_act = defaultdict(list)
    for r in char_rows:
        v = build_vec_from_xlsx(r)
        char_vecs_by_act[r.get('Activity', '?')].append(v)
    char_all = [(build_vec_from_xlsx(r), r.get('Activity', '?')) for r in char_rows]

    def hd_tuple(a, b):
        return sum(1 for x, y in zip(a, b) if norm_res(x) != norm_res(y))

    def nearest_hd(query_vec, target_vecs):
        if not target_vecs:
            return 999
        return min(hd_tuple(query_vec, t) for t in target_vecs)

    for pool_name, pos, target, target_acts in marker_pools:
        pool = [r for r in pv_rows if pv_res(r, pos) == target]
        if not pool:
            continue
        pool_vecs = [build_vec_from_pv(r) for r in pool]

        # Unique vectors
        vec_cnt = Counter(pool_vecs)

        # Kingdom breakdown
        k_cnt = Counter(tax.get(r['accession'], {}).get('kingdom', '?') for r in pool)

        print(f'\n{pool_name} pool ({pos}={target}):')
        print(f'  Structures: {len(pool)}, unique vectors: {len(vec_cnt)}')
        print(f'  Kingdom: {dict(k_cnt.most_common(4))}')
        top_v = vec_cnt.most_common(1)[0]
        print(f'  Top vector: {"-".join(top_v[0])} (n={top_v[1]})')

        # HD to nearest characterized target activity
        if target_acts is not None:
            if isinstance(target_acts, str):
                target_acts = [target_acts]
            tgt_vecs = []
            for ta in target_acts:
                tgt_vecs.extend(char_vecs_by_act.get(ta, []))

            hd_to_tgt = [nearest_hd(v, tgt_vecs) for v in pool_vecs]
            # HD to any characterized
            hd_to_any = [nearest_hd(v, [cv for cv, _ in char_all]) for v in pool_vecs]

            max_hd = max(max(hd_to_tgt), 5)
            print(f'  HD to nearest {target_acts} characterized:')
            cum = 0
            for d in range(0, min(max_hd + 1, 12)):
                n_at = sum(1 for h in hd_to_tgt if h == d)
                cum += n_at
                pct = cum / len(pool)
                print(f'    HD={d}: {n_at:5} cumulative={cum:5} ({pct:.1%})')
            print(f'  HD to nearest ANY characterized:')
            cum = 0
            for d in range(0, 8):
                n_at = sum(1 for h in hd_to_any if h == d)
                cum += n_at
                pct = cum / len(pool)
                if n_at > 0 or d <= 2:
                    print(f'    HD={d}: {n_at:5} cumulative={cum:5} ({pct:.1%})')
        # Position-level composition
        print(f'  Position composition:')
        for p in POSITIONS:
            cnt = Counter(pv_res(r, p) for r in pool)
            top = cnt.most_common(3)
            top_str = ', '.join(f'{aa}:{n}' for aa, n in top)
            print(f'    {p:12}: {top_str}')

    sep_print('DONE')
    print(f'\nAnalysis complete. {total_pv} structures, {len(char_rows)} characterized entries.')


if __name__ == '__main__':
    main()
