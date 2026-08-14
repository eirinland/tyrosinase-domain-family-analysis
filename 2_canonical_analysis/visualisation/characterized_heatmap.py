"""
Heatmap of characterized PPO vectors: 84 rows × 11 columns, grouped by activity.
Each cell shows the residue identity, colored by whether it matches the most
common residue at that position across the full dataset.
"""

import csv
from pathlib import Path
from collections import Counter

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap, to_rgb
import openpyxl

BASE = Path(__file__).parent.parent
VECTORS_CSV = BASE / 'position_vectors.csv'
CHAR_XLSX = Path('/cluster/work/projects/nn1003k/eirin/bioinf/characterized_PPOs.xlsx')

POSITIONS = ['Gly46', 'Phe65', 'Trp68', 'Glu195', 'Asn205',
             'Arg209', 'Val218', 'Ala221', 'Phe227', 'His230', 'thioether']

ACTIVITY_ORDER = ['TYR', 'CaOx', 'AUS', 'oAPO', 'oMP', 'DHICA ox', 'DCT', 'hemocyanin']

ACTIVITY_COLOR = {
    'TYR':        '#4C8CC2',
    'CaOx':       '#87B5D1',
    'AUS':        '#9ECCC7',
    'oAPO':       '#A6D1A6',
    'oMP':        '#6BAB76',
    'DHICA ox':   '#F0CFB0',
    'DCT':        '#D9A68A',
    'hemocyanin': '#B59EC7',
}

POS_COL_MAP = {
    'Gly46': 4, 'Phe65': 5, 'Trp68': 6, 'Glu195': 7, 'Asn205': 8,
    'Arg209': 9, 'Val218': 10, 'Ala221': 11, 'Phe227': 12, 'His230': 13,
    'thioether': 14,
}


def main():
    # Load full dataset to get baseline frequencies per position
    pos_col_csv = {
        'Gly46': 6, 'Phe65': 14, 'Trp68': 18, 'Glu195': 26, 'Asn205': 34,
        'Arg209': 42, 'Val218': 46, 'Ala221': 50, 'Phe227': 54, 'His230': 58,
        'thioether': 66,
    }
    baseline = {p: Counter() for p in POSITIONS}
    with open(VECTORS_CSV) as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            for pos in POSITIONS:
                val = row[pos_col_csv[pos]].replace('*', '')
                if pos == 'thioether':
                    val = 'C' if val in ('C', 'C*') else '-'
                baseline[pos][val] += 1

    most_common = {p: baseline[p].most_common(1)[0][0] for p in POSITIONS}
    print("Most common residue per position:")
    for p in POSITIONS:
        mc, ct = baseline[p].most_common(1)[0]
        total = sum(baseline[p].values())
        print(f"  {p}: {mc} ({100*ct/total:.1f}%)")

    # Load characterized set
    wb = openpyxl.load_workbook(CHAR_XLSX, read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc = str(row[0]).strip()
        act = str(row[1]).strip()
        vec_parts = []
        for pos in POSITIONS:
            col_idx = header.index(pos)
            val = str(row[col_idx]).strip().replace('*', '') if row[col_idx] else '?'
            if pos == 'thioether':
                val = 'C' if val in ('C', 'C*') else '-'
            vec_parts.append(val)
        entries.append({'acc': acc, 'activity': act, 'vector': vec_parts})
    wb.close()

    # Sort by activity order, then accession
    act_rank = {a: i for i, a in enumerate(ACTIVITY_ORDER)}
    entries.sort(key=lambda e: (act_rank.get(e['activity'], 99), e['acc']))

    n_entries = len(entries)
    n_pos = len(POSITIONS)

    # Build unique residue alphabet per position for coloring
    all_residues = set()
    for e in entries:
        all_residues.update(e['vector'])
    all_residues = sorted(all_residues)

    # Color scheme: matching consensus = white, different = colored by residue
    # Use a categorical approach: assign each unique residue a numeric code
    residue_to_idx = {r: i for i, r in enumerate(all_residues)}

    # Build matrix
    matrix = np.zeros((n_entries, n_pos))
    labels = []
    for i, e in enumerate(entries):
        for j, val in enumerate(e['vector']):
            if val == most_common[POSITIONS[j]]:
                matrix[i, j] = 0  # consensus
            else:
                matrix[i, j] = 1  # variant

    fig, (ax_bar, ax_heat) = plt.subplots(
        1, 2, figsize=(10, 16),
        gridspec_kw={'width_ratios': [0.4, 10], 'wspace': 0.02})

    # Activity color bar on left
    for i, e in enumerate(entries):
        color = ACTIVITY_COLOR.get(e['activity'], '#999999')
        ax_bar.barh(n_entries - 1 - i, 1, height=0.9, color=color, edgecolor='none')
    ax_bar.set_xlim(0, 1)
    ax_bar.set_ylim(-0.5, n_entries - 0.5)
    ax_bar.set_xticks([])
    ax_bar.set_yticks([])
    ax_bar.axis('off')

    # Heatmap
    cmap = ListedColormap(['#f0f0f0', '#fee08b'])
    ax_heat.imshow(matrix[::-1], aspect='auto', cmap=cmap, interpolation='nearest',
                   vmin=0, vmax=1)

    # Add text labels in each cell
    for i, e in enumerate(entries):
        for j, val in enumerate(e['vector']):
            color = '#333333' if matrix[i, j] == 1 else '#aaaaaa'
            fontweight = 'bold' if matrix[i, j] == 1 else 'normal'
            ax_heat.text(j, n_entries - 1 - i, val,
                        ha='center', va='center', fontsize=6,
                        color=color, fontweight=fontweight)

    # Y-axis: accession labels
    ax_heat.set_yticks(range(n_entries))
    ax_heat.set_yticklabels([e['acc'] for e in entries][::-1], fontsize=5.5, fontfamily='monospace')

    # X-axis: position names
    ax_heat.set_xticks(range(n_pos))
    ax_heat.set_xticklabels(POSITIONS, rotation=45, ha='right', fontsize=8)

    # Grid lines between activities
    current_act = entries[0]['activity']
    for i, e in enumerate(entries):
        if e['activity'] != current_act:
            ax_heat.axhline(y=n_entries - i - 0.5, color='black', linewidth=0.8)
            current_act = e['activity']

    # Activity labels on far right
    act_positions = {}
    for i, e in enumerate(entries):
        act = e['activity']
        if act not in act_positions:
            act_positions[act] = []
        act_positions[act].append(n_entries - 1 - i)

    for act, positions in act_positions.items():
        mid = np.mean(positions)
        ax_heat.text(n_pos + 0.3, mid, act, fontsize=7,
                    va='center', ha='left',
                    color=ACTIVITY_COLOR.get(act, '#333333'),
                    fontweight='bold')

    # Column header: consensus residue
    for j, pos in enumerate(POSITIONS):
        ax_heat.text(j, n_entries + 0.3, most_common[pos],
                    ha='center', va='bottom', fontsize=7,
                    color='#666666', fontstyle='italic')

    ax_heat.set_xlim(-0.5, n_pos - 0.5)
    ax_heat.set_ylim(-0.5, n_entries - 0.5 + 1)

    out = Path(__file__).parent / 'characterized_heatmap'
    fig.savefig(f'{out}.pdf', dpi=300, bbox_inches='tight')
    fig.savefig(f'{out}.png', dpi=200, bbox_inches='tight')
    plt.close()
    print(f"Saved: {out}.pdf / .png")


if __name__ == '__main__':
    main()
