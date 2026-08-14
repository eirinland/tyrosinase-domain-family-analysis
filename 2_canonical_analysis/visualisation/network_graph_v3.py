"""
Network graph of active-site vectors, v3 (decluttered).
- Higher MIN_SIZE to reduce small-count noise
- HD=1 edges only, ~ NOT treated as wildcard (real residue distance)
- Stronger repulsion for cleaner layout
- Characterized vectors always included regardless of member count
"""

import csv
from pathlib import Path
from collections import Counter

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import networkx as nx
import openpyxl

BASE = Path(__file__).parent.parent
VECTORS_CSV = BASE / 'position_vectors.csv'
CHAR_XLSX = Path('/cluster/work/projects/nn1003k/eirin/bioinf/characterized_PPOs.xlsx')

MIN_SIZE = 10
HAMMING_THRESHOLD = 1

ACTIVITY_COLOR = {
    'TYR':        '#4C8CC2',
    'CaOx':       '#87B5D1',
    'AUS':        '#9ECCC7',
    'oAPO':       '#A6D1A6',
    'oMP':        '#6BAB76',
    'DHICA ox':   '#F0CFB0',
    'DCT':        '#D9A68A',
    'hemocyanin': '#B59EC7',
}

POOL_COLOR = {
    'oAPO':       '#D3E8D3',
    'oMP':        '#B5D5BA',
    'DCT/DHICA':  '#F5E4D5',
    'hemocyanin': '#DACFE5',
}

COL = {
    'acc': 0, 'gly46': 6, 'gly46_ss': 8, 'phe65': 14, 'trp68': 18,
    'glu195': 26, 'asn205': 34, 'arg209': 42, 'val218': 46, 'ala221': 50,
    'phe227': 54, 'his230': 58, 'thioether': 66, 'vector': 68,
}


def hamming_strict(v1, v2):
    """HD without wildcard — ~ counts as a real character."""
    p1 = v1.split('-')
    p2 = v2.split('-')
    return sum(1 for a, b in zip(p1, p2) if a != b)


def load_characterized():
    wb = openpyxl.load_workbook(CHAR_XLSX, read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    acc_to_act = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc_to_act[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()
    return acc_to_act


def main():
    char_map = load_characterized()

    vec_count = Counter()
    vec_activities = {}
    vec_pools = {}

    with open(VECTORS_CSV) as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            acc = row[COL['acc']]
            vec = row[COL['vector']].replace('*', '')
            vec_count[vec] += 1

            if acc in char_map:
                vec_activities.setdefault(vec, set()).add(char_map[acc])

            gly46 = row[COL['gly46']].replace('*', '')
            arg209 = row[COL['arg209']].replace('*', '')
            his230 = row[COL['his230']].replace('*', '')
            pools = vec_pools.setdefault(vec, set())
            if gly46 == 'N':
                pools.add('oAPO')
            if arg209 == 'Y':
                pools.add('oMP')
            if his230 == 'L':
                pools.add('DCT/DHICA')
            if gly46 == 'E':
                pools.add('hemocyanin')

    # Keep vectors with >= MIN_SIZE members OR that contain characterized entries
    keep = {}
    for v, c in vec_count.items():
        if c >= MIN_SIZE or v in vec_activities:
            keep[v] = c
    vectors = sorted(keep.keys())
    print(f"Vectors (>= {MIN_SIZE} members or characterized): {len(vectors)}")
    print(f"  of which characterized: {sum(1 for v in vectors if v in vec_activities)}")

    G = nx.Graph()
    for v in vectors:
        G.add_node(v, count=keep[v])

    n = len(vectors)
    edges = []
    for i in range(n):
        for j in range(i + 1, n):
            d = hamming_strict(vectors[i], vectors[j])
            if d <= HAMMING_THRESHOLD:
                G.add_edge(vectors[i], vectors[j], weight=2.0)
                edges.append((vectors[i], vectors[j]))
    print(f"Edges (strict HD <= {HAMMING_THRESHOLD}): {len(edges)}")

    components = list(nx.connected_components(G))
    print(f"Connected components: {len(components)}")
    sizes_comp = sorted([len(c) for c in components], reverse=True)
    print(f"Component sizes (top 10): {sizes_comp[:10]}")

    pos = nx.spring_layout(G, k=3.5, iterations=300, seed=42, weight='weight')

    sizes = np.array([keep[v] for v in vectors])
    log_sizes = np.log10(np.clip(sizes, 1, None))
    rng = max(log_sizes.max() - log_sizes.min(), 1)
    pt_sizes = 20 + (log_sizes - log_sizes.min()) / rng * 400

    # Assign colors
    colors = []
    for v in vectors:
        if v in vec_activities:
            acts = sorted(vec_activities[v])
            colors.append(ACTIVITY_COLOR.get(acts[0], '#555555'))
        elif vec_pools.get(v):
            pools = sorted(vec_pools[v])
            colors.append(POOL_COLOR.get(pools[0], '#dddddd'))
        else:
            colors.append('#dddddd')

    fig, ax = plt.subplots(figsize=(16, 13))

    # Edges
    nx.draw_networkx_edges(G, pos, edgelist=edges, ax=ax,
                           edge_color='#bbbbbb', width=0.4, alpha=0.4)

    # Layer 1: uncharacterized, non-pool
    plain_idx = [i for i, v in enumerate(vectors)
                 if v not in vec_activities and not vec_pools.get(v)]
    if plain_idx:
        xy = np.array([pos[vectors[i]] for i in plain_idx])
        ax.scatter(xy[:, 0], xy[:, 1],
                   s=[pt_sizes[i] for i in plain_idx],
                   c=[colors[i] for i in plain_idx],
                   edgecolors='#aaaaaa', linewidths=0.2, alpha=0.45, zorder=2)

    # Layer 2: pool members (not characterized)
    pool_idx = [i for i, v in enumerate(vectors)
                if v not in vec_activities and vec_pools.get(v)]
    if pool_idx:
        xy = np.array([pos[vectors[i]] for i in pool_idx])
        ax.scatter(xy[:, 0], xy[:, 1],
                   s=[pt_sizes[i] for i in pool_idx],
                   c=[colors[i] for i in pool_idx],
                   edgecolors='#777777', linewidths=0.5, alpha=0.7, zorder=3)

    # Layer 3: characterized (diamonds, larger, black edge)
    char_idx = [i for i, v in enumerate(vectors) if v in vec_activities]
    if char_idx:
        xy = np.array([pos[vectors[i]] for i in char_idx])
        ax.scatter(xy[:, 0], xy[:, 1],
                   s=[pt_sizes[i] * 2.5 for i in char_idx],
                   c=[colors[i] for i in char_idx],
                   edgecolors='black', linewidths=1.5, alpha=1.0, zorder=4,
                   marker='D')

    # Legend
    legend_elements = []
    for act in ['TYR', 'CaOx', 'AUS', 'oAPO', 'oMP', 'DHICA ox', 'DCT', 'hemocyanin']:
        if any(act in vec_activities.get(v, set()) for v in vectors):
            legend_elements.append(
                Line2D([0], [0], marker='D', color='w',
                       markerfacecolor=ACTIVITY_COLOR[act],
                       markeredgecolor='black', markersize=9,
                       label=f'{act} (characterized)'))

    legend_elements.append(Line2D([0], [0], marker='s', color='w', markersize=0, label=''))
    for pool_name in ['oAPO', 'oMP', 'DCT/DHICA', 'hemocyanin']:
        legend_elements.append(
            Line2D([0], [0], marker='o', color='w',
                   markerfacecolor=POOL_COLOR[pool_name],
                   markeredgecolor='#777777', markersize=8,
                   label=f'{pool_name} pool'))

    legend_elements.append(
        Line2D([0], [0], marker='o', color='w',
               markerfacecolor='#dddddd', markeredgecolor='#aaaaaa',
               markersize=7, label='Other'))

    legend_elements.append(Line2D([0], [0], marker='s', color='w', markersize=0, label=''))
    for s_example in [10, 100, 500]:
        if s_example <= sizes.max():
            ls = np.log10(s_example)
            ps = 20 + (ls - log_sizes.min()) / rng * 400
            legend_elements.append(
                Line2D([0], [0], marker='o', color='w', markerfacecolor='#aaaaaa',
                       markersize=np.sqrt(ps) * 0.5, label=f'n = {s_example}'))

    ax.legend(handles=legend_elements, loc='upper left', fontsize=9,
              framealpha=0.95, title='Activity / Pool / Size',
              title_fontsize=10, borderpad=1)

    ax.set_xticks([])
    ax.set_yticks([])
    ax.axis('off')

    out = Path(__file__).parent / 'network_graph_v3'
    fig.savefig(f'{out}.pdf', dpi=300, bbox_inches='tight')
    fig.savefig(f'{out}.png', dpi=200, bbox_inches='tight')
    plt.close()
    print(f"Saved: {out}.pdf / .png")


if __name__ == '__main__':
    main()
