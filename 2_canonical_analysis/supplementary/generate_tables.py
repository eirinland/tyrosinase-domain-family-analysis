"""
Generate all manuscript results: supplementary tables, marker-pool statistics,
controller-pair analysis, conservation stats, and Hamming distance coherence.

Inputs:
  - position_vectors.csv          (21,928 structures)
  - taxonomy_lookup.csv           (accession → kingdom, phylum, genus)
  - characterized_PPOs.xlsx       (84 characterized PPOs)
  - FoldSeek cluster_cluster.tsv  (structural clustering)

Outputs (in supplementary/):
  Tables:
    table_oAPO_helix.tsv, table_oAPO_loop.tsv, table_oMP_Arg209Y.tsv,
    table_oMP_Phe65L.tsv, table_DCT_DHICA.tsv, table_hemocyanin.tsv,
    table_Fusarium_Pseudomonas_convergence.tsv
  Stats:
    manuscript_results.txt  (all numbers reported in the manuscript)
"""

import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path(__file__).resolve().parent.parent
VECTORS_CSV = BASE / "position_vectors.csv"
TAXONOMY_CSV = BASE / "visualisation" / "taxonomy_lookup.csv"
CHAR_XLSX = Path("/cluster/work/projects/nn1003k/eirin/bioinf/characterized_PPOs.xlsx")
CLUSTER_TSV = Path(
    "/cluster/work/projects/nn1003k/eirin/bioinf/foldseek/"
    "plddt_trimmed_tm08_cc/with_step45_rejects/results/cluster_cluster.tsv"
)
OUTDIR = Path(__file__).resolve().parent

# Column indices in position_vectors.csv (0-based)
C = {
    "acc": 0, "gly46": 6, "gly46_ss": 8, "phe65": 14, "trp68": 18,
    "glu195": 26, "asn205": 34, "arg209": 42, "val218": 46,
    "ala221": 50, "phe227": 54, "his230": 58, "thioether": 66, "vector": 68,
}

POSITION_NAMES = ["Gly46", "Phe65", "Trp68", "Glu195", "Asn205",
                  "Arg209", "Val218", "Ala221", "Phe227", "His230", "thioether"]
POSITION_COLS = [C["gly46"], C["phe65"], C["trp68"], C["glu195"], C["asn205"],
                 C["arg209"], C["val218"], C["ala221"], C["phe227"], C["his230"],
                 C["thioether"]]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_taxonomy():
    tax = {}
    with open(TAXONOMY_CSV) as f:
        for row in csv.reader(f):
            if row[0] == "accession":
                continue
            tax[row[0]] = {"kingdom": row[2], "phylum": row[3], "genus": row[4]}
    return tax


def load_clusters():
    rep_members = defaultdict(list)
    with open(CLUSTER_TSV) as f:
        for line in f:
            rep, member = line.strip().split("\t")
            rep_members[rep.split("_taxID_")[0]].append(member.split("_taxID_")[0])
    clusters = sorted(rep_members, key=lambda r: len(rep_members[r]), reverse=True)
    acc_to_cluster = {}
    for i, rep in enumerate(clusters, 1):
        for mem in rep_members[rep]:
            acc_to_cluster[mem] = f"cluster_{i}"
    return acc_to_cluster


def load_characterized():
    """Return list of dicts with accession, activity, and per-position residues."""
    wb = openpyxl.load_workbook(CHAR_XLSX, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        e = {
            "accession": str(row[0]).strip(),
            "activity": str(row[1]).strip(),
            "vector": str(row[-1]).strip() if row[-1] else "",
        }
        # Positional residues are columns 4-13 (0-indexed) in the xlsx
        # Gly46, Phe65, Trp68, Glu195, Asn205, Arg209, Val218, Ala221, Phe227, His230
        for i, name in enumerate(["Gly46", "Phe65", "Trp68", "Glu195", "Asn205",
                                   "Arg209", "Val218", "Ala221", "Phe227", "His230"]):
            val = row[4 + i]
            e[name] = str(val).strip().replace("*", "") if val else "?"
        e["thioether"] = str(row[14]).strip().replace("*", "") if row[14] else "-"
        e["vector"] = e["vector"].replace("*", "")
        entries.append(e)
    wb.close()
    return entries


def load_all_rows():
    rows = []
    with open(VECTORS_CSV) as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            row[C["vector"]] = row[C["vector"]].replace("*", "")
            rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_vector(v):
    v = v.strip()
    if v.endswith("--"):
        parts = v[:-2].split("-")
        parts.append("-")
    else:
        parts = v.split("-")
    return parts


def strip_star(s):
    return s.rstrip("*")


def hamming(v1, v2):
    d = 0
    for i in range(min(len(v1), len(v2))):
        a, b = strip_star(v1[i]), strip_star(v2[i])
        if a == "~" or b == "~":
            continue
        if a != b:
            d += 1
    return d


def min_hd(vec_parsed, refs_parsed):
    return min(hamming(vec_parsed, r) for r in refs_parsed)


def tax_summary(accs, taxonomy):
    kings = Counter(taxonomy.get(a, {}).get("kingdom", "?") for a in accs)
    return kings


def fmt_tax(kings, n):
    lines = []
    for k, c in kings.most_common():
        lines.append(f"    {k}: {c} ({c/n*100:.1f}%)")
    return "\n".join(lines)


def write_tsv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(header)
        for row in rows:
            w.writerow(row)
    print(f"  {path.name}: {len(rows)} rows")


def get_tax(acc, taxonomy):
    t = taxonomy.get(acc, {"kingdom": "?", "phylum": "?", "genus": "?"})
    return t["kingdom"], t["phylum"], t["genus"]


def hd_block(pool_vectors, ref_vectors_raw, label):
    refs = [parse_vector(v) for v in ref_vectors_raw]
    hds = [min_hd(parse_vector(v), refs) for v in pool_vectors]
    n = len(hds)
    if n == 0:
        return f"\n  {label}: empty pool\n"
    ctr = Counter(hds)
    uv = len(set(pool_vectors))
    le1 = sum(v for k, v in ctr.items() if k <= 1)
    le2 = sum(v for k, v in ctr.items() if k <= 2)
    median = sorted(hds)[n // 2]
    lines = [f"\n  {label} (n={n}, {uv} unique vectors):"]
    cum = 0
    for d in sorted(ctr):
        cum += ctr[d]
        lines.append(f"    HD={d}: {ctr[d]:>5} ({ctr[d]/n*100:5.1f}%)  cumulative: {cum/n*100:5.1f}%")
    lines.append(f"    HD<=1: {le1/n*100:.1f}%   HD<=2: {le2/n*100:.1f}%   Median: {median}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Controller pair analysis
# ---------------------------------------------------------------------------
def controller_pair_analysis(char_entries, out):
    out.append("\n\nASN205+ARG209 CONTROLLER PAIR")
    out.append("=" * 50)

    pair_activities = defaultdict(lambda: defaultdict(list))
    for e in char_entries:
        pair = (e["Asn205"], e["Arg209"])
        pair_activities[pair][e["activity"]].append(e["accession"])

    unique_pairs = set()
    for e in char_entries:
        unique_pairs.add((e["Asn205"], e["Arg209"]))
    out.append(f"Unique Asn205+Arg209 pairs in characterized set: {len(unique_pairs)}")

    # Shared signatures
    shared = {p: acts for p, acts in pair_activities.items() if len(acts) > 1}
    out.append(f"\nShared pairs (same pair in multiple activities): {len(shared)}")
    for (n205, r209), acts in sorted(shared.items()):
        parts = []
        for act, accs in sorted(acts.items()):
            parts.append(f"{act}({len(accs)})")
        total = sum(len(a) for a in acts.values())
        out.append(f"  {n205}+{r209}: {' + '.join(parts)} = {total} entries")

    # For each shared pair, count how many of the 9 non-marker positions
    # perfectly separate the activities
    out.append("\nSeparating positions per shared pair (of 9 remaining positions):")
    sep_positions = ["Gly46", "Phe65", "Trp68", "Glu195", "Val218",
                     "Ala221", "Phe227", "His230", "thioether"]
    for (n205, r209), acts in sorted(shared.items()):
        act_list = sorted(acts.keys())
        if len(act_list) != 2:
            continue
        a1, a2 = act_list
        separating = []
        for pos in sep_positions:
            vals1 = set(e[pos] for e in char_entries
                        if e["Asn205"] == n205 and e["Arg209"] == r209 and e["activity"] == a1)
            vals2 = set(e[pos] for e in char_entries
                        if e["Asn205"] == n205 and e["Arg209"] == r209 and e["activity"] == a2)
            if not vals1.intersection(vals2):
                separating.append(pos)
        out.append(f"  {n205}+{r209} ({a1} vs {a2}): {len(separating)} of 9 — {', '.join(separating) if separating else 'none'}")


# ---------------------------------------------------------------------------
# Controller pair projection onto full dataset
# ---------------------------------------------------------------------------
def controller_pair_projection(char_entries, all_rows, out):
    out.append("\n\nCONTROLLER PAIR PROJECTION")
    out.append("=" * 50)

    # Build mapping: (Asn205, Arg209) → set of activities in characterized
    pair_acts = defaultdict(set)
    for e in char_entries:
        pair_acts[(e["Asn205"], e["Arg209"])].add(e["activity"])

    unambiguous = set()  # pairs mapping to exactly 1 activity
    shared_pairs = set()  # pairs mapping to 2+ activities
    for pair, acts in pair_acts.items():
        if len(acts) == 1:
            unambiguous.add(pair)
        else:
            shared_pairs.add(pair)

    n_unambig = 0
    n_shared = 0
    n_novel = 0
    unambig_by_act = defaultdict(int)

    for r in all_rows:
        pair = (r[C["asn205"]], r[C["arg209"]])
        if pair in unambiguous:
            n_unambig += 1
            act = next(iter(pair_acts[pair]))
            unambig_by_act[act] += 1
        elif pair in shared_pairs:
            n_shared += 1
        else:
            n_novel += 1

    total = len(all_rows)
    out.append(f"Total structures: {total}")
    out.append(f"Unambiguous (pair maps to 1 activity): {n_unambig} ({n_unambig/total*100:.1f}%)")
    out.append(f"Shared (pair maps to 2+ activities):   {n_shared} ({n_shared/total*100:.1f}%)")
    out.append(f"Novel (pair not in characterized set):  {n_novel} ({n_novel/total*100:.1f}%)")

    out.append(f"\nUnambiguous pool sizes:")
    for act, count in sorted(unambig_by_act.items(), key=lambda x: -x[1]):
        out.append(f"  {act}: {count}")

    return n_novel, n_unambig, n_shared


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading data...")
    taxonomy = load_taxonomy()
    clusters = load_clusters()
    all_rows = load_all_rows()
    char_entries = load_characterized()

    # Char vectors by activity
    char_vecs_by_act = defaultdict(set)
    for e in char_entries:
        if e["vector"]:
            char_vecs_by_act[e["activity"]].add(e["vector"])
    dct_dhica_refs = list(char_vecs_by_act.get("DCT", set()) | char_vecs_by_act.get("DHICA ox", set()))
    omp_refs = list(char_vecs_by_act.get("oMP", set()))
    oapo_refs = list(char_vecs_by_act.get("oAPO", set()))
    hemo_refs = list(char_vecs_by_act.get("hemocyanin", set()))

    out = []  # accumulate all output lines

    # ==================================================================
    # §1  BASIC COUNTS
    # ==================================================================
    out.append("BASIC COUNTS")
    out.append("=" * 50)
    out.append(f"Total structures: {len(all_rows)}")
    out.append(f"Characterized PPOs: {len(char_entries)}")

    act_counts = Counter(e["activity"] for e in char_entries)
    for act, n in act_counts.most_common():
        out.append(f"  {act}: {n}")

    all_accs = [r[C["acc"]] for r in all_rows]
    king_all = tax_summary(all_accs, taxonomy)
    out.append(f"\nOverall taxonomy:")
    out.append(fmt_tax(king_all, len(all_rows)))

    # Unique vectors
    all_vectors = [r[C["vector"]] for r in all_rows]
    unique_vecs = len(set(all_vectors))
    out.append(f"\nTotal unique vectors: {unique_vecs}")

    char_vectors_all = set(e["vector"] for e in char_entries if e["vector"])
    out.append(f"Characterized PPO unique vectors: {len(char_vectors_all)}")

    # ==================================================================
    # §2  GATEKEEPER (Val218)
    # ==================================================================
    out.append("\n\nGATEKEEPER (Val218)")
    out.append("=" * 50)

    # Val218 distribution across characterized activities
    val218_by_act = defaultdict(lambda: defaultdict(int))
    for e in char_entries:
        val218_by_act[e["Val218"]][e["activity"]] += 1

    out.append("Val218 across characterized activities:")
    for res in sorted(val218_by_act):
        acts = val218_by_act[res]
        parts = [f"{a}({c})" for a, c in sorted(acts.items(), key=lambda x: -x[1])]
        out.append(f"  Val218={res}: {' + '.join(parts)}")

    # TYR-specific gatekeeper stats
    tyrs = [e for e in char_entries if e["activity"] == "TYR"]
    n_tyr = len(tyrs)
    tyr_v218f = sum(1 for e in tyrs if e["Val218"] == "F")
    tyr_ff = sum(1 for e in tyrs if e["Val218"] == "F" and e["Trp68"] == "F")
    out.append(f"\nTYRs with Val218=F: {tyr_v218f} of {n_tyr} ({tyr_v218f/n_tyr*100:.1f}%)")
    out.append(f"TYRs with Trp68=F AND Val218=F: {tyr_ff} of {n_tyr} ({tyr_ff/n_tyr*100:.1f}%)")

    # CaOx gatekeeper
    caox = [e for e in char_entries if e["activity"] == "CaOx"]
    caox_ff = sum(1 for e in caox if e["Val218"] == "F" and e["Trp68"] == "F")
    out.append(f"CaOx with Trp68=F AND Val218=F: {caox_ff} of {len(caox)}")

    # Val218 diversity in TYR
    tyr_v218_residues = set(e["Val218"] for e in tyrs)
    out.append(f"Val218 diversity in TYR: {len(tyr_v218_residues)} residues ({', '.join(sorted(tyr_v218_residues))})")

    # ==================================================================
    # §3  CONTROLLER PAIR
    # ==================================================================
    controller_pair_analysis(char_entries, out)
    n_novel, n_unambig, n_shared = controller_pair_projection(char_entries, all_rows, out)

    # ==================================================================
    # §4  MARKER-DEFINED GROUPS
    # ==================================================================
    out.append("\n\nMARKER-DEFINED GROUPS")
    out.append("=" * 50)

    # Pre-compute sets
    arg209y_set = {r[C["acc"]] for r in all_rows if r[C["arg209"]] == "Y"}

    # Collectors for tables and HD
    oapo_helix_rows, oapo_loop_rows = [], []
    omp_arg209y_rows, omp_phe65l_rows = [], []
    dct_dhica_rows, hemo_rows, fus_pseudo_rows = [], [], []
    hd_omp, hd_dct, hd_oapo_full, hd_oapo_helix, hd_hemo = [], [], [], [], []

    # Pool membership tracking
    pool_membership = defaultdict(set)

    for r in all_rows:
        acc = r[C["acc"]]
        vec = r[C["vector"]]
        clust = clusters.get(acc, "?")
        kingdom, phylum, genus = get_tax(acc, taxonomy)
        gly46_ss = {"a": "helix", "c": "coil", "b": "sheet"}.get(r[C["gly46_ss"]], r[C["gly46_ss"]])

        if r[C["gly46"]] == "N":
            pool_membership[acc].add("oAPO")
            hd_oapo_full.append(vec)
            if r[C["gly46_ss"]] == "a":
                oapo_helix_rows.append([acc, clust, vec, gly46_ss, kingdom, phylum, genus])
                hd_oapo_helix.append(vec)
            else:
                oapo_loop_rows.append([acc, clust, vec, gly46_ss, kingdom, phylum, genus])

        if r[C["arg209"]] == "Y":
            pool_membership[acc].add("oMP")
            omp_arg209y_rows.append([acc, clust, vec, kingdom, phylum, genus])
            hd_omp.append(vec)

        if r[C["phe65"]] == "L":
            in_y = "yes" if acc in arg209y_set else "no"
            omp_phe65l_rows.append([acc, clust, vec, r[C["arg209"]], in_y, kingdom, phylum, genus])

        if r[C["his230"]] == "L":
            pool_membership[acc].add("DCT/DHICA")
            arg209 = r[C["arg209"]]
            val218 = r[C["val218"]]
            if arg209 == "L" and val218 == "T":
                subsig = "DHICA-like"
            elif arg209 == "S" and val218 == "P":
                subsig = "DCT-like"
            else:
                subsig = "other"
            dct_dhica_rows.append([acc, clust, vec, arg209, val218, subsig, kingdom, phylum, genus])
            hd_dct.append(vec)

        if r[C["gly46"]] == "E":
            pool_membership[acc].add("hemocyanin")
            hemo_rows.append([acc, clust, vec, gly46_ss, kingdom, phylum, genus])
            hd_hemo.append(vec)

        if (r[C["asn205"]] == "D" and r[C["arg209"]] == "G" and
                r[C["val218"]] == "N" and r[C["his230"]] == "Y"):
            fus_pseudo_rows.append([acc, clust, vec, kingdom, phylum, genus])

    # --- oAPO ---
    out.append("\noAPO — Gly46=N:")
    n_oapo = len(oapo_helix_rows) + len(oapo_loop_rows)
    out.append(f"  Total: {n_oapo}")
    out.append(f"  Helix context: {len(oapo_helix_rows)}")
    out.append(f"  Non-helix: {len(oapo_loop_rows)}")
    out.append(f"  Unique vectors (full): {len(set(hd_oapo_full))}")
    out.append(f"  Unique vectors (helix): {len(set(hd_oapo_helix))}")

    helix_accs = [r[0] for r in oapo_helix_rows]
    helix_tax = tax_summary(helix_accs, taxonomy)
    out.append(f"  Helix-context taxonomy:")
    out.append(fmt_tax(helix_tax, len(helix_accs)))

    # Diversity at key positions within oAPO pool
    oapo_all_accs = [r[0] for r in oapo_helix_rows + oapo_loop_rows]
    oapo_rows_all = [r for r in all_rows if r[C["gly46"]] == "N"]
    for pos_name, col in [("Arg209", C["arg209"]), ("Asn205", C["asn205"]), ("Val218", C["val218"])]:
        residues = set(r[col] for r in oapo_rows_all)
        out.append(f"  Distinct residues at {pos_name}: {len(residues)}")

    # B1VTI5 vector match count
    b1vti5_vec = "N-F-W-E-G-V-I-S-F-N--"
    b1vti5_count = sum(1 for r in all_rows if r[C["vector"]] == b1vti5_vec)
    out.append(f"  B1VTI5 vector ({b1vti5_vec}) matches: {b1vti5_count}")

    # D6RTB9 vector match count
    d6rtb9_vec = "N-F-W-E-G-Q-I-S-F-H--"
    d6rtb9_count = sum(1 for r in all_rows if r[C["vector"]] == d6rtb9_vec)
    out.append(f"  D6RTB9 vector ({d6rtb9_vec}) matches: {d6rtb9_count}")

    # Asn+Ile combination
    asn_ile = sum(1 for r in all_rows if r[C["gly46"]] == "N" and r[C["val218"]] == "I")
    out.append(f"  Gly46=N AND Val218=I combination: {asn_ile}")

    # --- oMP ---
    out.append(f"\noMP — Arg209=Y:")
    n_omp = len(omp_arg209y_rows)
    out.append(f"  Total: {n_omp} ({n_omp/len(all_rows)*100:.1f}%)")
    out.append(f"  Unique vectors: {len(set(hd_omp))}")

    omp_tax = tax_summary([r[0] for r in omp_arg209y_rows], taxonomy)
    fungi_omp = omp_tax.get("Fungi", 0)
    fungi_overall_pct = king_all.get("Fungi", 0) / len(all_rows) * 100
    enrichment = (fungi_omp / n_omp * 100) / fungi_overall_pct if fungi_overall_pct > 0 else 0
    out.append(f"  Taxonomy:")
    out.append(fmt_tax(omp_tax, n_omp))
    out.append(f"  Fungi enrichment: {enrichment:.1f}x")

    # Internal position stats
    omp_rows = [r for r in all_rows if r[C["arg209"]] == "Y"]
    gln_glu195 = sum(1 for r in omp_rows if r[C["glu195"]] == "Q")
    gly_asn205 = sum(1 for r in omp_rows if r[C["asn205"]] == "G")
    asn_asn205 = sum(1 for r in omp_rows if r[C["asn205"]] == "N")
    ser_ala221 = sum(1 for r in omp_rows if r[C["ala221"]] == "S")
    no_thio = sum(1 for r in omp_rows if r[C["thioether"]] == "-")
    out.append(f"  Internal positions:")
    out.append(f"    Gln at Glu195: {gln_glu195} ({gln_glu195/n_omp*100:.1f}%)")
    out.append(f"    Gly at Asn205: {gly_asn205} ({gly_asn205/n_omp*100:.1f}%)")
    out.append(f"    Asn at Asn205: {asn_asn205}")
    out.append(f"    Ser at Ala221: {ser_ala221} ({ser_ala221/n_omp*100:.1f}%)")
    out.append(f"    No thioether:  {no_thio} ({no_thio/n_omp*100:.1f}%)")

    # Phe65=L overlap
    n_phe65l = len(omp_phe65l_rows)
    phe65l_in_y = sum(1 for r in omp_phe65l_rows if r[4] == "yes")
    phe65l_not_y = n_phe65l - phe65l_in_y
    out.append(f"  Phe65=L total: {n_phe65l}")
    out.append(f"    Also Arg209=Y: {phe65l_in_y} ({phe65l_in_y/n_phe65l*100:.1f}%)")
    out.append(f"    Without Arg209=Y: {phe65l_not_y} ({phe65l_not_y/n_phe65l*100:.1f}%)")

    # Arg209 distribution in Phe65=L without Arg209=Y
    non_overlap_arg209 = Counter(r[3] for r in omp_phe65l_rows if r[4] == "no")
    out.append(f"    Arg209 in non-overlapping: {', '.join(f'{res}({c})' for res, c in non_overlap_arg209.most_common(6))}")

    # oMP markers in characterized set
    char_omp = [e for e in char_entries if e["activity"] == "oMP"]
    omp_with_y = sum(1 for e in char_omp if e["Arg209"] == "Y")
    omp_with_l = sum(1 for e in char_omp if e["Phe65"] == "L")
    omp_with_both = sum(1 for e in char_omp if e["Phe65"] == "L" and e["Arg209"] == "Y")
    out.append(f"  Characterized oMP markers:")
    out.append(f"    Arg209=Y: {omp_with_y}/{len(char_omp)}")
    out.append(f"    Phe65=L:  {omp_with_l}/{len(char_omp)}")
    out.append(f"    Both:     {omp_with_both}/{len(char_omp)}")
    escapes = [e["accession"] for e in char_omp if e["Arg209"] != "Y" and e["Phe65"] != "L"]
    if escapes:
        out.append(f"    Escapes both markers: {', '.join(escapes)}")

    # --- DCT/DHICA ---
    out.append(f"\nDCT/DHICA — His230=L:")
    n_dct = len(dct_dhica_rows)
    out.append(f"  Total: {n_dct} ({n_dct/len(all_rows)*100:.1f}%)")
    out.append(f"  Unique vectors: {len(set(hd_dct))}")

    dct_tax = tax_summary([r[0] for r in dct_dhica_rows], taxonomy)
    animals_dct = dct_tax.get("Animals", 0)
    animals_pct = king_all.get("Animals", 0) / len(all_rows) * 100
    dct_enrichment = (animals_dct / n_dct * 100) / animals_pct if animals_pct > 0 else 0
    out.append(f"  Taxonomy:")
    out.append(fmt_tax(dct_tax, n_dct))
    out.append(f"  Animals enrichment: {dct_enrichment:.2f}x")
    non_animal = n_dct - animals_dct
    out.append(f"  Non-animal: {non_animal}")

    # Verified: all characterized DCT+DHICA have His230=L
    char_dct_dhica = [e for e in char_entries if e["activity"] in ("DCT", "DHICA ox")]
    all_have_l = all(e["His230"] == "L" for e in char_dct_dhica)
    out.append(f"  All {len(char_dct_dhica)} characterized DCT+DHICA have His230=L: {all_have_l}")

    # Sub-signatures
    dhica_like = [r for r in dct_dhica_rows if r[5] == "DHICA-like"]
    dct_like = [r for r in dct_dhica_rows if r[5] == "DCT-like"]
    other_dct = [r for r in dct_dhica_rows if r[5] == "other"]
    out.append(f"  Sub-signatures:")
    out.append(f"    DHICA-like (Arg209=L, Val218=T): {len(dhica_like)} ({len(dhica_like)/n_dct*100:.1f}%)")
    dhica_animals = sum(1 for r in dhica_like if r[6] == "Animals")
    out.append(f"      Animals: {dhica_animals}/{len(dhica_like)} ({dhica_animals/len(dhica_like)*100:.1f}%)")
    out.append(f"    DCT-like (Arg209=S, Val218=P): {len(dct_like)} ({len(dct_like)/n_dct*100:.1f}%)")
    dct_animals = sum(1 for r in dct_like if r[6] == "Animals")
    out.append(f"      Animals: {dct_animals}/{len(dct_like)} ({dct_animals/len(dct_like)*100:.1f}%)")
    other_uvec = len(set(r[2] for r in other_dct))
    out.append(f"    Other: {len(other_dct)} ({len(other_dct)/n_dct*100:.1f}%), {other_uvec} unique vectors")
    other_tax = tax_summary([r[0] for r in other_dct], taxonomy)
    out.append(f"    Other taxonomy:")
    out.append(fmt_tax(other_tax, len(other_dct)))

    # Exact vector matches to characterized
    dct_dhica_all = [r for r in all_rows if r[C["his230"]] == "L"]
    exact_counts = {}
    for label, vec in [("DHICA vector (~-F-W-E-N-L-T-S-F-L--)", "~-F-W-E-N-L-T-S-F-L--"),
                       ("P55028 vector (V-F-W-E-N-L-T-S-F-L--)", "V-F-W-E-N-L-T-S-F-L--"),
                       ("DCT vector (~-F-W-E-N-S-P-A-F-L--)", "~-F-W-E-N-S-P-A-F-L--")]:
        count = sum(1 for r in dct_dhica_all if r[C["vector"]] == vec)
        exact_counts[vec] = count
        out.append(f"    {label}: {count}")
    total_exact = sum(exact_counts.values())
    out.append(f"    Total exact matches: {total_exact}")

    # --- Hemocyanin ---
    out.append(f"\nHemocyanin — Gly46=E:")
    n_hemo = len(hemo_rows)
    out.append(f"  Total: {n_hemo}")
    out.append(f"  Unique vectors: {len(set(hd_hemo))}")

    hemo_ss = Counter(r[3] for r in hemo_rows)
    out.append(f"  Secondary structure: helix={hemo_ss.get('helix',0)}, coil={hemo_ss.get('coil',0)}, sheet={hemo_ss.get('sheet',0)}")

    # All characterized hemocyanins — check ss context
    char_hemo = [e for e in char_entries if e["activity"] == "hemocyanin"]
    out.append(f"  Characterized hemocyanins: {len(char_hemo)}")
    for e in char_hemo:
        match = [r for r in all_rows if r[C["acc"]] == e["accession"]]
        if match:
            ss = {"a": "helix", "c": "coil", "b": "sheet"}.get(match[0][C["gly46_ss"]], "?")
            out.append(f"    {e['accession']}: {ss}, Val218={e['Val218']}")

    hemo_tax = tax_summary([r[0] for r in hemo_rows], taxonomy)
    out.append(f"  Taxonomy:")
    out.append(fmt_tax(hemo_tax, n_hemo))

    # Val218 in hemocyanin pool
    hemo_all = [r for r in all_rows if r[C["gly46"]] == "E"]
    hemo_v218f = sum(1 for r in hemo_all if r[C["val218"]] == "F")
    out.append(f"  Val218=F in pool: {hemo_v218f} ({hemo_v218f/n_hemo*100:.1f}%)")

    # --- Fusarium/Pseudomonas ---
    out.append(f"\n119-structure His230=Y core (Asn205=D, Arg209=G, Val218=N, His230=Y):")
    out.append(f"  Count: {len(fus_pseudo_rows)}")
    fp_tax = tax_summary([r[0] for r in fus_pseudo_rows], taxonomy)
    out.append(f"  Taxonomy:")
    out.append(fmt_tax(fp_tax, len(fus_pseudo_rows)))
    fp_genus = Counter(r[5] for r in fus_pseudo_rows)
    out.append(f"  Top genera: {', '.join(f'{g}({c})' for g, c in fp_genus.most_common(5))}")

    # Thioether in this subset
    fp_all = [r for r in all_rows if r[C["asn205"]] == "D" and r[C["arg209"]] == "G"
              and r[C["val218"]] == "N" and r[C["his230"]] == "Y"]
    thio_c = sum(1 for r in fp_all if r[C["thioether"]] == "C")
    thio_cs = sum(1 for r in fp_all if r[C["thioether"]] == "C*")
    out.append(f"  Thioether: C={thio_c}, C*={thio_cs}")

    # ==================================================================
    # §5  POOL TOTALS
    # ==================================================================
    out.append("\n\nPOOL TOTALS")
    out.append("=" * 50)
    in_any = sum(1 for v in pool_membership.values() if len(v) >= 1)
    in_multi = sum(1 for v in pool_membership.values() if len(v) >= 2)
    remaining = len(all_rows) - in_any
    out.append(f"Structures in any pool: {in_any} ({in_any/len(all_rows)*100:.1f}%)")
    out.append(f"In 2+ pools: {in_multi}")
    out.append(f"Remaining: {remaining}")

    # ==================================================================
    # §6  CONSERVATION STATS
    # ==================================================================
    out.append("\n\nCONSERVATION STATS")
    out.append("=" * 50)

    his230_h = sum(1 for r in all_rows if r[C["his230"]] == "H")
    out.append(f"His230=H conservation: {his230_h}/{len(all_rows)} ({his230_h/len(all_rows)*100:.1f}%)")
    non_q = [r for r in all_rows if taxonomy.get(r[C["acc"]], {}).get("kingdom", "?") != "?"]
    his230_h_nq = sum(1 for r in non_q if r[C["his230"]] == "H")
    out.append(f"  Excluding ? kingdom: {his230_h_nq}/{len(non_q)} ({his230_h_nq/len(non_q)*100:.1f}%)")

    ser_ala221_all = sum(1 for r in all_rows if r[C["ala221"]] == "S")
    out.append(f"\nSer at Ala221: {ser_ala221_all}/{len(all_rows)} ({ser_ala221_all/len(all_rows)*100:.1f}%)")

    # Which activities have Ser at Ala221?
    out.append("  Per characterized activity:")
    for act in sorted(set(e["activity"] for e in char_entries)):
        act_entries = [e for e in char_entries if e["activity"] == act]
        has_ser = sum(1 for e in act_entries if e["Ala221"] == "S")
        out.append(f"    {act}: {has_ser}/{len(act_entries)}")

    # His230 variants
    his230_n = sum(1 for r in all_rows if r[C["his230"]] == "N")
    his230_y = sum(1 for r in all_rows if r[C["his230"]] == "Y")
    out.append(f"\nHis230=N: {his230_n}")
    out.append(f"His230=Y: {his230_y}")

    # Trp68=Y
    trp68y = [r for r in all_rows if r[C["trp68"]] == "Y"]
    out.append(f"\nTrp68=Y: {len(trp68y)}")
    trp68y_tax = tax_summary([r[C["acc"]] for r in trp68y], taxonomy)
    out.append(fmt_tax(trp68y_tax, len(trp68y)))

    # ==================================================================
    # §7  NOVEL FRACTION
    # ==================================================================
    out.append("\n\nNOVEL FRACTION")
    out.append("=" * 50)

    # Novel = Asn205+Arg209 pair not seen in characterized set
    char_pairs = set()
    for e in char_entries:
        char_pairs.add((e["Asn205"], e["Arg209"]))

    novel_rows = [r for r in all_rows if (r[C["asn205"]], r[C["arg209"]]) not in char_pairs]
    n_nov = len(novel_rows)
    out.append(f"Novel fraction: {n_nov} ({n_nov/len(all_rows)*100:.1f}%)")
    nov_tax = tax_summary([r[C["acc"]] for r in novel_rows], taxonomy)
    out.append(f"Taxonomy:")
    out.append(fmt_tax(nov_tax, n_nov))

    non_asn_205 = sum(1 for r in novel_rows if r[C["asn205"]] != "N")
    out.append(f"Non-Asn at Asn205: {non_asn_205}/{n_nov} ({non_asn_205/n_nov*100:.1f}%)")

    # % non-Asn at Asn205 in characterized TYRs
    tyr_non_asn = sum(1 for e in tyrs if e["Asn205"] != "N")
    out.append(f"Non-Asn at Asn205 in characterized TYRs: {tyr_non_asn}/{n_tyr} ({tyr_non_asn/n_tyr*100:.1f}%)")

    # ==================================================================
    # §8  SECONDARY STRUCTURE AT GLY46
    # ==================================================================
    out.append("\n\nSECONDARY STRUCTURE AT GLY46")
    out.append("=" * 50)
    ss_counts = Counter(r[C["gly46_ss"]] for r in all_rows)
    for ss_code, label in [("a", "helix"), ("c", "coil"), ("b", "sheet"), ("NA", "NA (all-? vectors)")]:
        c = ss_counts.get(ss_code, 0)
        out.append(f"  {label}: {c} ({c/len(all_rows)*100:.1f}%)")

    # Helix enrichment per Gly46 residue
    gly46_residues = Counter(r[C["gly46"]] for r in all_rows)
    out.append(f"\n  Helix fraction per Gly46 residue:")
    for res, total in gly46_residues.most_common():
        helix_count = sum(1 for r in all_rows if r[C["gly46"]] == res and r[C["gly46_ss"]] == "a")
        out.append(f"    {res}: {helix_count}/{total} ({helix_count/total*100:.1f}% helix)")

    # ==================================================================
    # §9  HAMMING DISTANCE STATS
    # ==================================================================
    out.append("\n\nHAMMING DISTANCE TO CHARACTERIZED VECTORS")
    out.append("=" * 50)
    out.append("~ at Gly46 (loop context) treated as wildcard — skipped in HD.")
    out.append(f"\nReference vectors: oMP={len(omp_refs)}, DCT/DHICA={len(dct_dhica_refs)}, "
               f"oAPO={len(oapo_refs)}, hemocyanin={len(hemo_refs)}")

    out.append(hd_block(hd_omp, omp_refs, "oMP (Arg209=Y)"))
    out.append(hd_block(hd_dct, dct_dhica_refs, "DCT/DHICA (His230=L)"))
    out.append(hd_block(hd_oapo_full, oapo_refs, "oAPO full (Gly46=N)"))
    out.append(hd_block(hd_oapo_helix, oapo_refs, "oAPO helix-only"))
    out.append(hd_block(hd_hemo, hemo_refs, "hemocyanin (Gly46=E)"))

    # ~ at Gly46 counts per pool
    tilde_omp = sum(1 for v in hd_omp if parse_vector(v)[0] == "~")
    tilde_dct = sum(1 for v in hd_dct if parse_vector(v)[0] == "~")
    tilde_oapo = sum(1 for v in hd_oapo_full if parse_vector(v)[0] == "~")
    out.append(f"\n  ~ at Gly46 in pools:")
    out.append(f"    oMP: {tilde_omp}/{len(hd_omp)} ({tilde_omp/len(hd_omp)*100:.1f}%)")
    out.append(f"    DCT/DHICA: {tilde_dct}/{len(hd_dct)} ({tilde_dct/len(hd_dct)*100:.1f}%)")
    out.append(f"    oAPO: {tilde_oapo}/{len(hd_oapo_full)} ({tilde_oapo/len(hd_oapo_full)*100:.1f}%)")

    # ==================================================================
    # Write results file
    # ==================================================================
    results_file = OUTDIR / "manuscript_results.txt"
    with open(results_file, "w") as f:
        f.write("\n".join(out) + "\n")
    print(f"\n  {results_file.name}: written")

    # ==================================================================
    # Write supplementary tables
    # ==================================================================
    print("\nWriting supplementary tables...")
    write_tsv(OUTDIR / "table_oAPO_helix.tsv",
              ["accession", "structural_cluster", "vector", "secondary_structure_Gly46",
               "kingdom", "phylum", "genus"],
              sorted(oapo_helix_rows))
    write_tsv(OUTDIR / "table_oAPO_loop.tsv",
              ["accession", "structural_cluster", "vector", "secondary_structure_Gly46",
               "kingdom", "phylum", "genus"],
              sorted(oapo_loop_rows))
    write_tsv(OUTDIR / "table_oMP_Arg209Y.tsv",
              ["accession", "structural_cluster", "vector", "kingdom", "phylum", "genus"],
              sorted(omp_arg209y_rows))
    write_tsv(OUTDIR / "table_oMP_Phe65L.tsv",
              ["accession", "structural_cluster", "vector", "Arg209_residue",
               "in_Arg209Y_group", "kingdom", "phylum", "genus"],
              sorted(omp_phe65l_rows))
    write_tsv(OUTDIR / "table_DCT_DHICA.tsv",
              ["accession", "structural_cluster", "vector", "Arg209_residue",
               "Val218_residue", "sub_signature", "kingdom", "phylum", "genus"],
              sorted(dct_dhica_rows))
    write_tsv(OUTDIR / "table_hemocyanin.tsv",
              ["accession", "structural_cluster", "vector", "secondary_structure_Gly46",
               "kingdom", "phylum", "genus"],
              sorted(hemo_rows))
    write_tsv(OUTDIR / "table_Fusarium_Pseudomonas_convergence.tsv",
              ["accession", "structural_cluster", "vector", "kingdom", "phylum", "genus"],
              sorted(fus_pseudo_rows))

    print("\nDone.")


if __name__ == "__main__":
    main()
